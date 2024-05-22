import csv
import os
from random import randint
data = []
sort_order = ["Type", "Rarity", "Element", "Name", "Quantity", "Sell"]  # Custom
sort_order = ["Type", "Rarity", "Element", "Name", "Quantity"]
all_keys = ["Ref", "Type", "Rarity", "Element", "Name", "Quantity", "Sell"]  # Must contain all header keys
cat_keys = ["Type", "Rarity", "Element"]  # categorical
num_keys = ["Ref", "Quantity", "Sell", "Name"]  # numerical. (Name is encoded... decoded?)
filename = "myData.csv"
fileout = "myDataSorted.csv"
fileout2 = "myDataSorted.xlsx"
ignore_non_sorted = True


def base27encoder(myString):  # a to z (97-122) mapped to some integer in base10.
    # aaa = 0 and zzz = 25(26^2) + 25(26^1) + 25 = 17575
    # EXCEPT! aaa should not be 0.
    # so aaa = 1(27^2) + 1(27^1) + 1 = 757
    # and zzz = 26(27^2) + 26(27^1) + 26 = 19682
    # watch out for big numbers in other languages
    total = 0
    for i in range(len(myString)):
        char = 0
        if myString[i] != 0:
            char = ord(myString[i]) - 97 + 1  # only lower case letters, i'm not checking.
        total += char*(27**(len(myString) - i - 1))  # reversed order... oops!
    return total
# try print(base27encoder('aaa'))
# quit()


def create_data(size):
    gen_dict = {  # ideally, you should map each word to a number instead to sort.
        "Type":
            ["Material", "Tool", "Armor", "Food", "Meal", "Potion", "Material/Food", "Construct"],
        "Rarity":
            ["Common", "Rare", "Epic", "Legendary"],
        "Element":
            ["water", "air", "earth", "bacteria", "plant", "life", "fire", "energy", "wave", "gravity", "time", "space"]
    }

    def randname(length):
        name = ""
        for ci in range(length):
            name += chr(randint(97, 122))
        return name

    for i in range(size):
        data.append(
            {"Ref": i,
             "Type": gen_dict["Type"][randint(0, len(gen_dict["Type"])-1)],
             "Rarity": gen_dict["Rarity"][randint(0, len(gen_dict["Rarity"])-1)],
             "Element": gen_dict["Element"][randint(0, len(gen_dict["Element"]) - 1)],
             "Name": randname(10),
             "Quantity": randint(1, 101)
             }
        )
        if randint(0, 1) == 1:
            data[i-1]["Sell"] = randint(1, 10) * 50
    return data


# data = create_data(500)  # write data option
# if not os.path.isfile('./'+filename):
#     f = open(filename, 'x')
# with open(filename, 'w', newline='') as csvfile:
#     writer = csv.DictWriter(csvfile, fieldnames=all_keys)
#     writer.writeheader()
#     writer.writerows(data)

with open(filename, mode='r') as file:  # read data option
    csvFile = csv.reader(file)

    head = None
    first = True
    count = 0
    for lines in csvFile:
        if first:
            head = lines
            first = False
            continue
        data.append(dict(zip(head, lines)))

        for CLASS in sort_order:
            if data[count].get(CLASS) == '':  # remove empty
                data[count].pop(CLASS, None)
        count += 1


# for i in range(300):
#     print(i, chr(i))  # 33 ... 48... 65 (A) - 90 (Z) and 97 (a) - 122 (z) and beyond
# for x in data:
#     print(x)


def recur_sort(input_array, output_array, var_list, var_i):  # if index starts at 1, +1
    temp_output_array = []
    CLASS = var_list[var_i]  # [0] select one of the vars, in order
    for row in input_array:  # [6a] if the row does not have the class, it must have a stand-in
        if row.get(CLASS) is None:
            if CLASS in cat_keys:  # (additional)
                row[CLASS] = None
            elif CLASS in num_keys:
                row[CLASS] = '0'

    if CLASS in num_keys:  # (additional)
        if CLASS == "Name":
            input_array = sorted(input_array, key=lambda row: base27encoder(row[CLASS]))  # [3] Sort by that value
        else:
            input_array = sorted(input_array, key=lambda row: int(row[CLASS]))
    elif CLASS in cat_keys:
        input_array = sorted(input_array, key=lambda row: row[CLASS])  # [3] Sort by that value

    if var_i + 2 > len(var_list):  # [4] No more class. Ex: (4) + 2 > 5
        return input_array

    unique_list = []
    for row in input_array:  # [1] get unique values
        if row[CLASS] not in unique_list:
            unique_list.append(row[CLASS])

    for class0 in unique_list:  # for each unique value...
        low_array = []
        for row in input_array:  # [2] get all rows with that value
            if row[CLASS] == class0:
                low_array.append(row)

        sorted_array = recur_sort(low_array, output_array, var_list, var_i+1)
        for row5 in sorted_array:
            temp_output_array.append(row5)  # [5] avoid infinity loop.

    for row in input_array:  # [6b] remove stand-in
        if row.get(CLASS) is None:
            row.pop(CLASS, None)
    return temp_output_array

# Ex: input = [ [ index (ref), type, class, rarity, element, name, bar, quantity, sell] ... ]
output = recur_sort(data, [], sort_order, 0)
# for row in output:
#     print(row)
print("Match:", data == output)
print("Length:", len(data), len(output))

if not os.path.isfile('./'+fileout):  # write new file
    f = open(fileout, 'x')
with open(fileout, 'w', newline='') as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=all_keys)
    writer.writeheader()
    writer.writerows(output)

# https://stackoverflow.com/questions/17684610/python-convert-csv-to-xlsx
import pandas as pd
read_file = pd.read_csv(fileout)
read_file.to_excel(fileout2, index=None, header=True)  # pip3 install  openpyxl




# https://stackoverflow.com/questions/25408393/getting-individual-colors-from-a-color-map-in-matplotlib
# https://stackoverflow.com/questions/72337841/convert-rgba-tuple-to-hex-with-matplotlib
import matplotlib
cmap = matplotlib.cm.get_cmap('Spectral')
import xlsxwriter
# https://www.geeksforgeeks.org/how-to-add-colour-to-excel-cells-using-python/
import openpyxl
from openpyxl.styles import PatternFill
wb = openpyxl.load_workbook(fileout2)
ws = wb['Sheet1']


# To map values, if you are too lazy
unique_list_all = {}


for CLASS in all_keys:  # check each class
    unique_list = []
    lowest = float('inf')
    highest = float('-inf')
    for row in output:  # per row
        if CLASS in num_keys:
            if row[CLASS] is None:
                continue
            val = row[CLASS]
            if not str.isdigit(val):  # name
                val = base27encoder(val)
            if int(val) < lowest:  # int because me no like float
                lowest = int(val)
            elif int(val) > highest:
                highest = int(val)
        elif CLASS in cat_keys:
            if row[CLASS] not in unique_list:
                if row[CLASS] is None:
                    continue
                unique_list.append(row[CLASS])

    if CLASS in num_keys:
        unique_list = [lowest, highest]
    unique_list.sort()
    unique_list_all[CLASS] = unique_list
print(unique_list_all)


row_no = 1
for row in output:
    class_no = 0
    for CLASS in all_keys:
        if ignore_non_sorted and CLASS not in sort_order:  # skip
            class_no += 1  # fix
            continue

        if row.get(CLASS) is not None:  # https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
            # for numerical, fit 0 to max number
            rgba = cmap(0.01)
            if CLASS in num_keys and row[CLASS] is not None:
                val = row[CLASS]
                if not str.isdigit(val):  # name
                    val = base27encoder(val)
                else:
                    val = int(val)
                # let 0 be the lowest and 1 be the highest (it's a line)
                ls = unique_list_all[CLASS]
                a = ls[0]  # ???, 0
                b = ls[1]  # ???, 1
                m = (0-1)/(1.0*(a-b))  # c = 0  # + m*a
                # print(CLASS, m*val)
                rgba = cmap(m*val)
            elif CLASS in cat_keys and row[CLASS] is not None:
                val = row[CLASS]
                rgba = cmap((unique_list_all[CLASS].index(val)+0) / len(unique_list_all[CLASS]))
                # if there's two classes, 0/2 and 1/2... add 0.5?
                # print(CLASS, unique_list_all[CLASS].index(val), len(unique_list_all[CLASS]) )

            myColor = matplotlib.colors.rgb2hex(rgba, keep_alpha=True)
            # print(myColor)
            color = myColor.upper()[7:9] + myColor.upper()[1:7]  # rgba to argb
            filler = PatternFill(patternType='solid', fgColor=color)
            # all_keys determine letter: A to Z, AA ... rows determine index (1 = header)
            cell_id = xlsxwriter.utility.xl_col_to_name(class_no) + str(row_no+1)
            ws[cell_id].fill = filler
        class_no += 1

    row_no += 1
    # break

wb.save(fileout2)
